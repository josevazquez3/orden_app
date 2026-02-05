"""
Módulo para cargar datos desde archivos Excel y Word
"""

import os
from openpyxl import load_workbook
from docx import Document
from typing import List, Tuple


class FileLoader:
    """Maneja la carga de datos desde archivos"""
    
    @staticmethod
    def cargar_desde_excel(ruta_archivo: str) -> List[Tuple[str, str]]:
        """
        Carga temas desde un archivo Excel
        
        El archivo debe tener las siguientes columnas:
        - Columna A: Descripción del tema
        - Columna B: Categoría (opcional)
        
        Returns:
            Lista de tuplas (descripcion, categoria)
        """
        if not os.path.exists(ruta_archivo):
            raise FileNotFoundError(f"El archivo {ruta_archivo} no existe")
        
        temas = []
        
        try:
            libro = load_workbook(ruta_archivo)
            hoja = libro.active
            
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                if fila[0]:  # Si hay contenido en la columna A
                    descripcion = str(fila[0]).strip()
                    categoria = str(fila[1]).strip() if fila[1] else None
                    
                    if descripcion:
                        temas.append((descripcion, categoria))
            
            libro.close()
            
        except Exception as e:
            raise Exception(f"Error al leer archivo Excel: {str(e)}")
        
        return temas
    
    @staticmethod
    def cargar_desde_word(ruta_archivo: str) -> List[Tuple[str, str]]:
        """
        Carga temas desde un documento Word
        
        Cada párrafo se interpreta como un tema.
        Si el párrafo contiene " - " se divide en descripción y categoría.
        
        Formato esperado:
        - Tema 1 - Categoría A
        - Tema 2 - Categoría B
        - Tema 3 (sin categoría)
        
        Returns:
            Lista de tuplas (descripcion, categoria)
        """
        if not os.path.exists(ruta_archivo):
            raise FileNotFoundError(f"El archivo {ruta_archivo} no existe")
        
        temas = []
        
        try:
            doc = Document(ruta_archivo)
            
            for parrafo in doc.paragraphs:
                texto = parrafo.text.strip()
                
                # Saltar párrafos vacíos
                if not texto:
                    continue
                
                # Saltar líneas que parecen títulos (todas mayúsculas, muy largas)
                if texto.isupper() and len(texto) > 50:
                    continue
                
                # Dividir por " - " si existe
                if ' - ' in texto:
                    descripcion, categoria = texto.split(' - ', 1)
                    descripcion = descripcion.strip()
                    categoria = categoria.strip()
                else:
                    descripcion = texto
                    categoria = None
                
                if descripcion:
                    temas.append((descripcion, categoria))
            
        except Exception as e:
            raise Exception(f"Error al leer documento Word: {str(e)}")
        
        return temas
    
    @staticmethod
    def cargar_desde_archivo(ruta_archivo: str) -> List[Tuple[str, str]]:
        """
        Detecta el tipo de archivo y carga los temas
        
        Args:
            ruta_archivo: Ruta al archivo (Excel o Word)
        
        Returns:
            Lista de tuplas (descripcion, categoria)
        
        Raises:
            ValueError: Si el formato de archivo no es soportado
        """
        if not os.path.exists(ruta_archivo):
            raise FileNotFoundError(f"El archivo {ruta_archivo} no existe")
        
        extension = os.path.splitext(ruta_archivo)[1].lower()
        
        if extension in ['.xlsx', '.xls']:
            return FileLoader.cargar_desde_excel(ruta_archivo)
        elif extension in ['.docx', '.doc']:
            return FileLoader.cargar_desde_word(ruta_archivo)
        else:
            raise ValueError(f"Formato de archivo no soportado: {extension}")
