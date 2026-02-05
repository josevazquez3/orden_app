"""
Controlador Principal
Conecta el modelo con las vistas y maneja la lógica de negocio
"""

from models.database import Database
from views.main_view import VentanaPrincipal
from views.dialogs import (DialogoTema, DialogoDelegado, DialogoSeleccionTema,
                           VentanaVistaPrevia, VentanaHistorialTema)
from utils.document_generator import DocumentGenerator
from tkinter import messagebox, END
import os


class MainController:
    """Controlador principal de la aplicación"""
    
    def __init__(self):
        # Inicializar modelo
        self.db = Database()
        self.db.cargar_datos_iniciales()
        
        # Inicializar vista
        self.view = VentanaPrincipal()
        
        # Inicializar generador de documentos
        self.doc_generator = DocumentGenerator()
        
        # Variables de estado
        self.orden_actual = []  # Lista de temas en el orden del día
        
        # Conectar eventos
        self._conectar_eventos()
        
        # Cargar datos iniciales
        self._cargar_datos_iniciales()
        
        print("✓ Controlador inicializado correctamente")
    
    def _conectar_eventos(self):
        """Conecta los botones de la vista con los métodos del controlador"""
        
        # === TAB REUNIÓN ===
        self.view.btn_agregar_tema.config(command=self._agregar_tema_orden)
        self.view.btn_subir.config(command=self._subir_tema)
        self.view.btn_bajar.config(command=self._bajar_tema)
        self.view.btn_eliminar.config(command=self._eliminar_tema_orden)
        self.view.btn_editar_delegado.config(command=self._editar_delegado_reunion)
        self.view.btn_subir_delegado.config(command=self._subir_delegado)
        self.view.btn_bajar_delegado.config(command=self._bajar_delegado)
        self.view.btn_vista_previa.config(command=self._mostrar_vista_previa)
        self.view.btn_generar_pdf.config(command=self._generar_pdf)
        self.view.btn_generar_doc.config(command=self._generar_doc)
        
        # === TAB TEMAS ===
        self.view.btn_nuevo_tema.config(command=self._nuevo_tema)
        self.view.btn_cargar_masivo.config(command=self._cargar_temas_desde_excel)
        self.view.btn_modificar_tema.config(command=self._modificar_tema)
        self.view.btn_eliminar_tema.config(command=self._eliminar_tema)
        self.view.btn_historial_tema.config(command=self._ver_historial_tema)
        self.view.btn_exportar_excel.config(command=self._exportar_temas_excel)
        self.view.btn_exportar_pdf.config(command=self._exportar_temas_pdf)
        self.view.btn_actualizar_temas.config(command=self._actualizar_lista_temas)
        
        # === TAB DELEGADOS ===
        self.view.btn_nuevo_delegado.config(command=self._nuevo_delegado)
        self.view.btn_modificar_delegado.config(command=self._modificar_delegado)
        self.view.btn_eliminar_delegado.config(command=self._eliminar_delegado)
        self.view.btn_actualizar_delegados.config(command=self._actualizar_lista_delegados)
        
        # === TAB HISTORIAL ===
        self.view.btn_actualizar_historial.config(command=self._actualizar_historial)
        self.view.btn_buscar_historial.config(command=self._buscar_historial)
        self.view.btn_limpiar_busqueda.config(command=self._limpiar_busqueda_historial)
        self.view.btn_exportar_historial_excel.config(command=self._exportar_historial_excel)
        self.view.btn_exportar_historial_pdf.config(command=self._exportar_historial_pdf)
        self.view.btn_borrar_historial.config(command=self._borrar_reuniones_seleccionadas)
    
    def _cargar_datos_iniciales(self):
        """Carga datos iniciales en las vistas"""
        self._actualizar_delegados_reunion()
        self._actualizar_combos_firmas()
        self._actualizar_lista_temas()
        self._actualizar_lista_delegados()
        self._actualizar_historial()
    
    # ==================== TAB REUNIÓN ====================
    
    def _actualizar_delegados_reunion(self):
        """Actualiza la tabla de delegados en el tab de reunión"""
        # Limpiar tabla
        for item in self.view.tree_delegados.get_children():
            self.view.tree_delegados.delete(item)
        
        # Cargar delegados titulares
        delegados = self.db.obtener_delegados(solo_titulares=True)
        for d in delegados:
            self.view.tree_delegados.insert('', 'end', values=(
                d['titulo'],
                d['nombre'],
                d['apellido'],
                d['distrito']
            ), tags=(d['id'],))
    
    def _actualizar_combos_firmas(self):
        """Actualiza los combobox de firmas (Presidente y Secretario)"""
        delegados = self.db.obtener_delegados(solo_titulares=True)
        nombres = [f"{d['titulo']} {d['nombre']} {d['apellido']}" for d in delegados]
        
        self.view.combo_presidente.config(values=nombres)
        self.view.combo_secretario.config(values=nombres)
        
        # Seleccionar por defecto Presidente (TUCCI) y Secretario (DUNOGENT)
        for i, d in enumerate(delegados):
            if 'TUCCI' in d['apellido']:
                self.view.combo_presidente.current(i)
            if 'DUNOGENT' in d['apellido']:
                self.view.combo_secretario.current(i)
    
    def _agregar_tema_orden(self):
        """Muestra diálogo para agregar tema al orden del día"""
        temas = self.db.obtener_temas()
        
        if not temas:
            messagebox.showinfo(
                "Información",
                "No hay temas disponibles.\nPor favor cree algunos temas primero en la pestaña 'Gestión de Temas'."
            )
            return
        
        # Mostrar diálogo de selección
        dialogo = DialogoSeleccionTema(self.view, temas)
        self.view.wait_window(dialogo)
        
        if dialogo.resultado:
            tema_id = int(dialogo.resultado)
            tema = self.db.obtener_tema(tema_id)
            
            # Agregar al orden actual
            numero_orden = len(self.orden_actual) + 1
            self.orden_actual.append({
                'tema_id': tema_id,
                'numero_orden': numero_orden,
                'descripcion': tema['descripcion']
            })
            
            self._actualizar_listbox_orden()
    
    def _subir_tema(self):
        """Sube el tema seleccionado en el orden"""
        seleccion = self.view.listbox_orden.curselection()
        
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un tema para subir")
            return
        
        indice = seleccion[0]
        
        if indice == 0:
            return  # Ya está al inicio
        
        # Intercambiar con el anterior
        self.orden_actual[indice], self.orden_actual[indice-1] = \
            self.orden_actual[indice-1], self.orden_actual[indice]
        
        # Actualizar números de orden
        for i, tema in enumerate(self.orden_actual):
            tema['numero_orden'] = i + 1
        
        self._actualizar_listbox_orden()
        
        # Mantener selección
        self.view.listbox_orden.selection_clear(0, 'end')
        self.view.listbox_orden.selection_set(indice-1)
    
    def _bajar_tema(self):
        """Baja el tema seleccionado en el orden"""
        seleccion = self.view.listbox_orden.curselection()
        
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un tema para bajar")
            return
        
        indice = seleccion[0]
        
        if indice == len(self.orden_actual) - 1:
            return  # Ya está al final
        
        # Intercambiar con el siguiente
        self.orden_actual[indice], self.orden_actual[indice+1] = \
            self.orden_actual[indice+1], self.orden_actual[indice]
        
        # Actualizar números de orden
        for i, tema in enumerate(self.orden_actual):
            tema['numero_orden'] = i + 1
        
        self._actualizar_listbox_orden()
        
        # Mantener selección
        self.view.listbox_orden.selection_clear(0, 'end')
        self.view.listbox_orden.selection_set(indice+1)
    
    def _eliminar_tema_orden(self):
        """Elimina el tema seleccionado del orden del día"""
        seleccion = self.view.listbox_orden.curselection()
        
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un tema para eliminar")
            return
        
        indice = seleccion[0]
        self.orden_actual.pop(indice)
        
        # Renumerar
        for i, tema in enumerate(self.orden_actual):
            tema['numero_orden'] = i + 1
        
        self._actualizar_listbox_orden()
    
    def _actualizar_listbox_orden(self):
        """Actualiza el listbox del orden del día"""
        self.view.listbox_orden.delete(0, 'end')
        
        for tema in self.orden_actual:
            texto = f"{tema['numero_orden']}.- {tema['descripcion']}"
            self.view.listbox_orden.insert('end', texto)
    
    def _editar_delegado_reunion(self):
        """Edita el delegado seleccionado en la tabla de reunión"""
        seleccion = self.view.tree_delegados.selection()
        
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un delegado para editar")
            return
        
        item = seleccion[0]
        delegado_id = int(self.view.tree_delegados.item(item, 'tags')[0])
        
        # Obtener el delegado
        delegados = self.db.obtener_delegados(solo_activos=False)
        delegado = next((d for d in delegados if d['id'] == delegado_id), None)
        
        if not delegado:
            return
        
        # Mostrar diálogo
        dialogo = DialogoDelegado(self.view, "Modificar Delegado", delegado)
        self.view.wait_window(dialogo)
        
        if dialogo.resultado:
            resultado = dialogo.resultado
            self.db.modificar_delegado(
                delegado_id,
                resultado['titulo'],
                resultado['nombre'],
                resultado['apellido'],
                resultado['distrito'],
                resultado['titular']
            )
            
            self._actualizar_delegados_reunion()
            self._actualizar_combos_firmas()
            messagebox.showinfo("Éxito", "Delegado modificado correctamente")
    
    def _subir_delegado(self):
        """Sube el delegado seleccionado en la lista"""
        seleccion = self.view.tree_delegados.selection()
        
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un delegado para subir")
            return
        
        # Obtener lista actual de delegados del treeview
        items = list(self.view.tree_delegados.get_children())
        indice_actual = items.index(seleccion[0])
        
        if indice_actual == 0:
            return  # Ya está al principio
        
        # Intercambiar con el anterior
        item_actual = seleccion[0]
        item_anterior = items[indice_actual - 1]
        
        # Obtener valores
        valores_actual = self.view.tree_delegados.item(item_actual, 'values')
        valores_anterior = self.view.tree_delegados.item(item_anterior, 'values')
        
        # Intercambiar en el treeview
        self.view.tree_delegados.item(item_anterior, values=valores_actual)
        self.view.tree_delegados.item(item_actual, values=valores_anterior)
        
        # Seleccionar el que se movió
        self.view.tree_delegados.selection_set(item_anterior)
    
    def _bajar_delegado(self):
        """Baja el delegado seleccionado en la lista"""
        seleccion = self.view.tree_delegados.selection()
        
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un delegado para bajar")
            return
        
        # Obtener lista actual de delegados del treeview
        items = list(self.view.tree_delegados.get_children())
        indice_actual = items.index(seleccion[0])
        
        if indice_actual == len(items) - 1:
            return  # Ya está al final
        
        # Intercambiar con el siguiente
        item_actual = seleccion[0]
        item_siguiente = items[indice_actual + 1]
        
        # Obtener valores
        valores_actual = self.view.tree_delegados.item(item_actual, 'values')
        valores_siguiente = self.view.tree_delegados.item(item_siguiente, 'values')
        
        # Intercambiar en el treeview
        self.view.tree_delegados.item(item_siguiente, values=valores_actual)
        self.view.tree_delegados.item(item_actual, values=valores_siguiente)
        
        # Seleccionar el que se movió
        self.view.tree_delegados.selection_set(item_siguiente)
    
    def _mostrar_vista_previa(self):
        """Muestra vista previa del documento"""
        if not self.orden_actual:
            messagebox.showwarning(
                "Advertencia",
                "Agregue al menos un tema al orden del día"
            )
            return
        
        datos = self._recopilar_datos_reunion()
        contenido = self.doc_generator.generar_texto_vista_previa(datos)
        
        VentanaVistaPrevia(self.view, contenido)
    
    def _generar_pdf(self):
        """Genera el documento PDF"""
        if not self.orden_actual:
            messagebox.showwarning(
                "Advertencia",
                "Agregue al menos un tema al orden del día"
            )
            return
        
        try:
            datos = self._recopilar_datos_reunion()
            archivo = self.doc_generator.generar_pdf(datos)
            
            # Guardar en base de datos
            self._guardar_reunion(datos)
            
            messagebox.showinfo(
                "Éxito",
                f"PDF generado correctamente:\n{archivo}"
            )
            
            # Intentar abrir el archivo
            try:
                os.system(f'xdg-open "{archivo}"')
            except:
                pass
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar PDF:\n{str(e)}")
    
    def _generar_doc(self):
        """Genera el documento DOCX"""
        if not self.orden_actual:
            messagebox.showwarning(
                "Advertencia",
                "Agregue al menos un tema al orden del día"
            )
            return
        
        try:
            datos = self._recopilar_datos_reunion()
            archivo = self.doc_generator.generar_docx(datos)
            
            # Guardar en base de datos
            self._guardar_reunion(datos)
            
            messagebox.showinfo(
                "Éxito",
                f"Documento generado correctamente:\n{archivo}"
            )
            
            # Intentar abrir el archivo
            try:
                os.system(f'xdg-open "{archivo}"')
            except:
                pass
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar documento:\n{str(e)}")
    
    def _recopilar_datos_reunion(self):
        """Recopila todos los datos de la reunión"""
        datos = {
            'fecha': self.view.entry_fecha.get(),
            'hora': self.view.entry_hora.get(),
            'lugar': self.view.entry_lugar.get(),
            'sede': self.view.entry_sede.get() or '',  # Asegurar que nunca sea None
            'tipo': self.view.combo_tipo.get(),
            'plataforma': self.view.plataforma.get() or '',  # Plataforma para reuniones virtuales
            'delegados': [],
            'orden_dia': self.orden_actual,
            'presidente': self.view.combo_presidente.get(),
            'secretario': self.view.combo_secretario.get(),
            'texto_encabezado': self.view.texto_encabezado.get(),
            'subtitulo_encabezado': self.view.subtitulo_encabezado.get(),
            'tamaño_titulo': self.view.tamaño_titulo.get(),
            'fuente_titulo': self.view.fuente_titulo.get(),
            'negrita_titulo': self.view.negrita_titulo.get(),
            'negrita_subtitulo': self.view.negrita_subtitulo.get(),
            'imagen_logo': self.view.imagen_path,
            'ancho_logo': self.view.tamaño_logo_ancho.get(),
            'alto_logo': self.view.tamaño_logo_alto.get(),
            'ancho_logo_docx': self.view.tamaño_logo_docx.get()
        }
        
        # Obtener delegados
        for item in self.view.tree_delegados.get_children():
            values = self.view.tree_delegados.item(item, 'values')
            datos['delegados'].append({
                'titulo': values[0],
                'nombre': values[1],
                'apellido': values[2],
                'distrito': values[3]
            })
        
        return datos
    
    def _guardar_reunion(self, datos):
        """Guarda la reunión en la base de datos"""
        # Crear reunión
        reunion_id = self.db.agregar_reunion(
            datos['fecha'],
            datos['hora'],
            datos['lugar'],
            datos['sede'],
            datos['tipo']
        )
        
        # Guardar orden del día
        for tema in datos['orden_dia']:
            self.db.agregar_tema_orden_dia(
                reunion_id,
                tema['tema_id'],
                tema['numero_orden']
            )
        
        # Guardar firmas
        delegados = self.db.obtener_delegados()
        presidente_id = None
        secretario_id = None
        
        for d in delegados:
            nombre_completo = f"{d['titulo']} {d['nombre']} {d['apellido']}"
            if nombre_completo == datos['presidente']:
                presidente_id = d['id']
            if nombre_completo == datos['secretario']:
                secretario_id = d['id']
        
        if presidente_id and secretario_id:
            self.db.guardar_firmas(reunion_id, presidente_id, secretario_id)
        
        self._actualizar_historial()
    
    # ==================== TAB TEMAS ====================
    
    def _actualizar_lista_temas(self):
        """Actualiza la lista de temas"""
        # Limpiar tabla
        for item in self.view.tree_temas.get_children():
            self.view.tree_temas.delete(item)
        
        # Cargar temas
        temas = self.db.obtener_temas(solo_activos=True)
        for tema in temas:
            stats = self.db.obtener_estadisticas_tema(tema['id'])
            estado = "Activo" if tema['activo'] else "Inactivo"
            
            self.view.tree_temas.insert('', 'end', values=(
                tema['id'],
                tema['descripcion'],
                tema['categoria'] or '-',
                stats['cantidad_usos'],
                estado
            ), tags=(tema['id'],))
    
    def _nuevo_tema(self):
        """Crea un nuevo tema"""
        dialogo = DialogoTema(self.view, "Nuevo Tema")
        self.view.wait_window(dialogo)
        
        if dialogo.resultado:
            resultado = dialogo.resultado
            self.db.agregar_tema(resultado['descripcion'], resultado['categoria'])
            self._actualizar_lista_temas()
            messagebox.showinfo("Éxito", "Tema creado correctamente")
    
    def _modificar_tema(self):
        """Modifica el tema seleccionado"""
        seleccion = self.view.tree_temas.selection()
        
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un tema para modificar")
            return

        if len(seleccion) > 1:
            messagebox.showwarning("Advertencia", "Seleccione un solo tema para modificar")
            return
        
        item = seleccion[0]
        tema_id = int(self.view.tree_temas.item(item, 'tags')[0])
        tema = self.db.obtener_tema(tema_id)
        
        dialogo = DialogoTema(self.view, "Modificar Tema", tema)
        self.view.wait_window(dialogo)
        
        if dialogo.resultado:
            resultado = dialogo.resultado
            self.db.modificar_tema(tema_id, resultado['descripcion'], resultado['categoria'])
            self._actualizar_lista_temas()
            messagebox.showinfo("Éxito", "Tema modificado correctamente")
    
    def _eliminar_tema(self):
        """Elimina el tema seleccionado"""
        seleccion = self.view.tree_temas.selection()

        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione al menos un tema para eliminar")
            return

        cantidad = len(seleccion)
        if not messagebox.askyesno("Confirmar eliminación", f"¿Está seguro de eliminar {cantidad} tema(s)?"):
            return

        borrados = 0
        errores = 0
        for item in seleccion:
            try:
                tema_id = int(self.view.tree_temas.item(item, 'tags')[0])
                if self.db.eliminar_tema(tema_id):
                    borrados += 1
                else:
                    errores += 1
            except Exception as e:
                print(f"Error al eliminar tema {item}: {e}")
                errores += 1

        self._actualizar_lista_temas()

        if errores == 0:
            messagebox.showinfo("Éxito", f"Se eliminaron {borrados} tema(s) correctamente")
        else:
            messagebox.showwarning("Eliminación Parcial", f"Se eliminaron {borrados} tema(s). {errores} no pudieron ser eliminados.")
    
    def _ver_historial_tema(self):
        """Muestra el historial del tema seleccionado"""
        seleccion = self.view.tree_temas.selection()
        
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un tema para ver su historial")
            return
        
        item = seleccion[0]
        tema_id = int(self.view.tree_temas.item(item, 'tags')[0])
        tema = self.db.obtener_tema(tema_id)
        historial = self.db.obtener_historial_tema(tema_id)
        stats = self.db.obtener_estadisticas_tema(tema_id)
        
        VentanaHistorialTema(self.view, tema, historial, stats)
    
    # ==================== TAB DELEGADOS ====================
    
    def _actualizar_lista_delegados(self):
        """Actualiza la lista de delegados"""
        # Limpiar tabla
        for item in self.view.tree_delegados_lista.get_children():
            self.view.tree_delegados_lista.delete(item)
        
        # Cargar delegados
        delegados = self.db.obtener_delegados(solo_activos=True, solo_titulares=False)
        for delegado in delegados:
            tipo = "Titular" if delegado['titular'] else "Suplente"
            
            self.view.tree_delegados_lista.insert('', 'end', values=(
                delegado['id'],
                delegado['titulo'],
                delegado['nombre'],
                delegado['apellido'],
                delegado['distrito'],
                tipo
            ), tags=(delegado['id'],))
    
    def _nuevo_delegado(self):
        """Crea un nuevo delegado"""
        dialogo = DialogoDelegado(self.view, "Nuevo Delegado")
        self.view.wait_window(dialogo)
        
        if dialogo.resultado:
            resultado = dialogo.resultado
            self.db.agregar_delegado(
                resultado['titulo'],
                resultado['nombre'],
                resultado['apellido'],
                resultado['distrito'],
                resultado['titular']
            )
            self._actualizar_lista_delegados()
            self._actualizar_delegados_reunion()
            self._actualizar_combos_firmas()
            messagebox.showinfo("Éxito", "Delegado creado correctamente")
    
    def _modificar_delegado(self):
        """Modifica el delegado seleccionado"""
        seleccion = self.view.tree_delegados_lista.selection()
        
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un delegado para modificar")
            return
        
        item = seleccion[0]
        delegado_id = int(self.view.tree_delegados_lista.item(item, 'values')[0])
        
        # Obtener el delegado
        delegados = self.db.obtener_delegados(solo_activos=False)
        delegado = next((d for d in delegados if d['id'] == delegado_id), None)
        
        if not delegado:
            return
        
        dialogo = DialogoDelegado(self.view, "Modificar Delegado", delegado)
        self.view.wait_window(dialogo)
        
        if dialogo.resultado:
            resultado = dialogo.resultado
            self.db.modificar_delegado(
                delegado_id,
                resultado['titulo'],
                resultado['nombre'],
                resultado['apellido'],
                resultado['distrito'],
                resultado['titular']
            )
            self._actualizar_lista_delegados()
            self._actualizar_delegados_reunion()
            self._actualizar_combos_firmas()
            messagebox.showinfo("Éxito", "Delegado modificado correctamente")
    
    def _eliminar_delegado(self):
        """Elimina el delegado seleccionado"""
        seleccion = self.view.tree_delegados_lista.selection()
        
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un delegado para eliminar")
            return
        
        if not messagebox.askyesno("Confirmar", "¿Está seguro de eliminar este delegado?"):
            return
        
        item = seleccion[0]
        delegado_id = int(self.view.tree_delegados_lista.item(item, 'values')[0])
        self.db.eliminar_delegado(delegado_id)
        self._actualizar_lista_delegados()
        self._actualizar_delegados_reunion()
        self._actualizar_combos_firmas()
        messagebox.showinfo("Éxito", "Delegado eliminado correctamente")
    
    # ==================== TAB HISTORIAL ====================
    
    def _actualizar_historial(self):
        """Actualiza el historial de reuniones"""
        # Limpiar tabla
        for item in self.view.tree_historial.get_children():
            self.view.tree_historial.delete(item)
        
        # Por ahora no hay reuniones (implementar después)
        # Esta funcionalidad requiere agregar método obtener_reuniones() en database.py
    
    # ==================== IMPORTAR/EXPORTAR TEMAS ====================
    
    def _cargar_temas_desde_excel(self):
        """Importa temas desde un archivo Excel"""
        from tkinter import filedialog
        import openpyxl
        
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        
        if not archivo:
            return
        
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            contador = 0
            errores = []
            
            # Asumir que la primera fila es encabezado
            for fila in ws.iter_rows(min_row=2, values_only=True):
                if not fila[0]:  # Si la primera columna está vacía, saltar
                    continue
                
                descripcion = str(fila[0]).strip() if fila[0] else ""
                categoria = str(fila[1]).strip() if len(fila) > 1 and fila[1] else ""
                
                if descripcion:
                    try:
                        self.db.agregar_tema(descripcion, categoria)
                        contador += 1
                    except Exception as e:
                        errores.append(f"Error en fila: {descripcion} - {str(e)}")
            
            self._actualizar_lista_temas()
            
            mensaje = f"Se importaron {contador} temas exitosamente."
            if errores:
                mensaje += f"\n\nErrores:\n" + "\n".join(errores[:5])
            
            messagebox.showinfo("Importación Exitosa", mensaje)
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al importar archivo: {str(e)}")
    
    def _exportar_temas_excel(self):
        """Exporta los temas a un archivo Excel"""
        from tkinter import filedialog
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        archivo = filedialog.asksaveasfilename(
            title="Guardar archivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            initialfile=f"Temas_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if not archivo:
            return
        
        try:
            temas = self.db.obtener_temas()
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Temas"
            
            # Encabezados
            encabezados = ["Descripción", "Categoría", "Veces Usado", "Estado"]
            ws.append(encabezados)
            
            # Formatear encabezados
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos
            for tema in temas:
                ws.append([
                    tema['descripcion'],
                    tema['categoria'] or "",
                    tema['veces_usado'],
                    "Activo" if tema['activo'] else "Inactivo"
                ])
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            
            wb.save(archivo)
            messagebox.showinfo("Exportación Exitosa", f"Se exportaron {len(temas)} temas a:\n{archivo}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel: {str(e)}")
    
    def _exportar_temas_pdf(self):
        """Exporta los temas a un archivo PDF"""
        from tkinter import filedialog
        from datetime import datetime
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib import colors
        
        archivo = filedialog.asksaveasfilename(
            title="Guardar archivo PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")],
            initialfile=f"Temas_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
        
        if not archivo:
            return
        
        try:
            temas = self.db.obtener_temas()
            
            # Crear documento PDF
            doc = SimpleDocTemplate(
                archivo,
                pagesize=A4,
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.black,
                spaceAfter=20,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            story.append(Paragraph("LISTADO DE TEMAS", title_style))
            story.append(Spacer(1, 0.5*cm))
            
            # Tabla de temas
            datos_tabla = [["Descripción", "Categoría", "Usos", "Estado"]]
            for tema in temas:
                datos_tabla.append([
                    tema['descripcion'],
                    tema['categoria'] or "",
                    str(tema['veces_usado']),
                    "Activo" if tema['activo'] else "Inactivo"
                ])
            
            tabla = Table(datos_tabla, colWidths=[9*cm, 3*cm, 1.5*cm, 2*cm])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            story.append(tabla)
            
            doc.build(story)
            messagebox.showinfo("Exportación Exitosa", f"Se exportaron {len(temas)} temas a:\n{archivo}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a PDF: {str(e)}")
    
    # ==================== HISTORIAL - BÚSQUEDA Y EXPORTACIÓN ====================
    
    def _buscar_historial(self):
        """Busca reuniones por tema o fecha"""
        termino = self.view.entry_buscar_historial.get().strip()
        
        if not termino:
            messagebox.showwarning("Búsqueda", "Por favor ingrese un término de búsqueda")
            return
        
        self._actualizar_lista_historial(termino)
    
    def _limpiar_busqueda_historial(self):
        """Limpia la búsqueda y muestra todas las reuniones"""
        self.view.entry_buscar_historial.delete(0, END)
        self._actualizar_historial()
    
    def _actualizar_historial(self):
        """Actualiza el historial de reuniones"""
        self._actualizar_lista_historial()
    
    def _actualizar_lista_historial(self, termino_busqueda: str = None):
        """Actualiza la tabla de historial"""
        # Limpiar tabla
        for item in self.view.tree_historial.get_children():
            self.view.tree_historial.delete(item)
        
        # Obtener reuniones
        if termino_busqueda:
            reuniones = self.db.buscar_reuniones(termino_busqueda)
        else:
            reuniones = self.db.obtener_reuniones()
        
        # Agregar a tabla
        for reunion in reuniones:
            # Obtener temas de la reunión
            temas = self.db.obtener_temas_reunion(reunion['id'])
            
            # Construir texto con temas y sus contadores de uso
            temas_con_usos = []
            for t in temas:
                stats = self.db.obtener_estadisticas_tema(t['id'])
                cantidad_usos = stats['cantidad_usos']
                temas_con_usos.append(f"{t['numero_orden']}. {t['descripcion']} ({cantidad_usos})")
            
            temas_texto = ", ".join(temas_con_usos)
            
            self.view.tree_historial.insert('', 'end', values=(
                reunion['id'],
                reunion['fecha'],
                reunion['hora'],
                reunion['lugar'],
                reunion['tipo'],
                temas_texto if temas_texto else "Sin temas"
            ))
    
    def _exportar_historial_excel(self):
        """Exporta el historial a Excel"""
        from tkinter import filedialog
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        archivo = filedialog.asksaveasfilename(
            title="Guardar archivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            initialfile=f"Historial_Reuniones_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if not archivo:
            return
        
        try:
            reuniones = self.db.obtener_reuniones()
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Historial"
            
            # Encabezados
            encabezados = ["ID", "Fecha", "Hora", "Lugar", "Tipo", "Temas"]
            ws.append(encabezados)
            
            # Formatear encabezados
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos
            for reunion in reuniones:
                # Obtener temas de la reunión
                temas = self.db.obtener_temas_reunion(reunion['id'])
                
                # Construir texto con temas y sus contadores de uso
                temas_con_usos = []
                for t in temas:
                    stats = self.db.obtener_estadisticas_tema(t['id'])
                    cantidad_usos = stats['cantidad_usos']
                    temas_con_usos.append(f"{t['numero_orden']}. {t['descripcion']} ({cantidad_usos})")
                
                temas_texto = ", ".join(temas_con_usos)
                
                ws.append([
                    reunion['id'],
                    reunion['fecha'],
                    reunion['hora'],
                    reunion['lugar'],
                    reunion['tipo'],
                    temas_texto if temas_texto else "Sin temas"
                ])
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 50
            
            wb.save(archivo)
            messagebox.showinfo("Exportación Exitosa", f"Se exportaron {len(reuniones)} reuniones a:\n{archivo}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel: {str(e)}")
    
    def _exportar_historial_pdf(self):
        """Exporta el historial a PDF"""
        from tkinter import filedialog
        from datetime import datetime
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib import colors
        
        archivo = filedialog.asksaveasfilename(
            title="Guardar archivo PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")],
            initialfile=f"Historial_Reuniones_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
        
        if not archivo:
            return
        
        try:
            reuniones = self.db.obtener_reuniones()
            
            # Crear documento PDF
            doc = SimpleDocTemplate(
                archivo,
                pagesize=A4,
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.black,
                spaceAfter=20,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            story.append(Paragraph("HISTORIAL DE REUNIONES", title_style))
            story.append(Spacer(1, 0.5*cm))
            
            # Tabla de reuniones
            datos_tabla = [["ID", "Fecha", "Hora", "Lugar", "Tipo", "Temas"]]
            for reunion in reuniones:
                # Obtener temas de la reunión
                temas = self.db.obtener_temas_reunion(reunion['id'])
                
                # Construir texto con temas y sus contadores de uso
                temas_con_usos = []
                for t in temas:
                    stats = self.db.obtener_estadisticas_tema(t['id'])
                    cantidad_usos = stats['cantidad_usos']
                    temas_con_usos.append(f"{t['numero_orden']}. {t['descripcion']} ({cantidad_usos})")
                
                temas_texto = ", ".join(temas_con_usos)
                
                datos_tabla.append([
                    str(reunion['id']),
                    reunion['fecha'],
                    reunion['hora'],
                    reunion['lugar'],
                    reunion['tipo'],
                    temas_texto if temas_texto else "Sin temas"
                ])
            
            tabla = Table(datos_tabla, colWidths=[1*cm, 2*cm, 1.5*cm, 4*cm, 2*cm, 4*cm])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('WORDWRAP', (0, 1), (-1, -1), True),
            ]))
            
            story.append(tabla)
            
            doc.build(story)
            messagebox.showinfo("Exportación Exitosa", f"Se exportaron {len(reuniones)} reuniones a:\n{archivo}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a PDF: {str(e)}")
    
    def _borrar_reuniones_seleccionadas(self):
        """Borra las reuniones seleccionadas del historial"""
        # Obtener items seleccionados
        seleccionados = self.view.tree_historial.selection()
        
        if not seleccionados:
            messagebox.showwarning("Advertencia", "Seleccione al menos una reunión para borrar")
            return
        
        # Confirmar eliminación
        cantidad = len(seleccionados)
        respuesta = messagebox.askyesno(
            "Confirmar eliminación",
            f"¿Está seguro de que desea eliminar {cantidad} reunión(es)?\n\nEsta acción no se puede deshacer."
        )
        
        if not respuesta:
            return
        
        # Obtener IDs de las reuniones seleccionadas
        reuniones_a_borrar = []
        for item in seleccionados:
            # El primer valor en values es el ID
            reunion_id = int(self.view.tree_historial.item(item, 'values')[0])
            reuniones_a_borrar.append(reunion_id)
        
        # Borrar cada reunión
        borradas = 0
        errores = 0
        
        for reunion_id in reuniones_a_borrar:
            try:
                if self.db.eliminar_reunion(reunion_id):
                    borradas += 1
                else:
                    errores += 1
            except Exception as e:
                print(f"Error al borrar reunión {reunion_id}: {e}")
                errores += 1
        
        # Actualizar la lista
        self._actualizar_historial()
        
        # Mostrar resultado
        if errores == 0:
            messagebox.showinfo("Éxito", f"Se eliminaron {borradas} reunión(es) correctamente")
        else:
            messagebox.showwarning(
                "Eliminación Parcial",
                f"Se eliminaron {borradas} reunión(es)\n{errores} no pudieron ser eliminadas"
            )

    
    def run(self):
        """Ejecuta la aplicación"""
        self.view.mainloop()